# -*- coding: utf-8 -*-
"""Coloca los compromisos del comité anterior en la columna de la lámina de Diagnóstico.
Uso: compromisos.fill(pptx_bytes, excel_bytes) -> pptx_bytes (con las tablas puestas).
- Cada campaña: tabla #  |  Compromiso  |  ¿Cumplió?  |  Comentarios (¿Cumplió?/Comentarios editables).
- Campañas sin compromisos en el Excel: misma tabla, vacía."""
import io, re, zipfile, unicodedata
from xml.dom import minidom
from collections import OrderedDict
E=914400; NAVY="1F4E79"; GUINDA="6B150B"

def _norm(s): return unicodedata.normalize("NFKD",str(s or "")).encode("ascii","ignore").decode().upper()
def _esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
def _sid(c):
    c=_norm(c)
    if "BCP" in c: return "BCP"
    if ("SCOTIABANK" in c or "SBP" in c) and "CONVENIO" in c: return "SCONV"
    if ("SCOTIABANK" in c or "SBP" in c) and "TARJETA" in c: return "STARJ"
    if "BBVA" in c and "CONVENIO" in c: return "BCONV"
    if "DIGITAL" in c and ("BBVA" in c or "TMK" in c or "TLM" in c): return "BDIG"
    if "OUT" in c and ("BBVA" in c or "TMK" in c or "TLM" in c): return "BOUT"
    if "UNICEF" in c: return "UNICEF"
    if "SANNA" in c: return "SANNA"
    if "DINERS" in c or "DINNERS" in c: return "DINERS"
    if "BANBIF" in c: return "BANBIF"
    if "EFECTIVA" in c or "FONO" in c: return "EFECT"
    return None
def _brand(s): return NAVY if s in ("BCONV","BDIG","BOUT","UNICEF") else GUINDA

def _read(xlsx):
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(xlsx), data_only=True)
    for sh in wb.sheetnames:
        rows=list(wb[sh].iter_rows(values_only=True))
        hi=cc=cp=None
        for i,r in enumerate(rows[:12]):
            cells=[_norm(x) for x in r]
            if any("CAMPA" in c for c in cells) and any("COMPROMISO" in c for c in cells):
                hi=i
                for j,c in enumerate(cells):
                    if "CAMPA" in c and cc is None: cc=j
                    if "COMPROMISO" in c and cp is None: cp=j
                break
        if hi is None: continue
        out=OrderedDict()
        for r in rows[hi+1:]:
            camp=r[cc] if cc<len(r) else None; comp=r[cp] if cp<len(r) else None
            if camp and comp:
                s=_sid(camp)
                if s: out.setdefault(s,[]).append(str(comp).strip())
        return out
    return OrderedDict()

def _tc(txt, sz=1050, bold=False, color="2B2B2B", fill="FFFFFF", algn="l", ital=False):
    b=' b="1"' if bold else ''; it=' i="1"' if ital else ''
    run=(f'<a:r><a:rPr lang="es-PE" sz="{sz}"{b}{it} dirty="0"><a:solidFill><a:srgbClr val="{color}"/></a:solidFill></a:rPr><a:t>{_esc(txt)}</a:t></a:r>'
         if txt!="" else '<a:endParaRPr lang="es-PE"/>')
    ln=''.join(f'<a:ln{s} w="9525"><a:solidFill><a:srgbClr val="D9DEE4"/></a:solidFill></a:ln{s}>' for s in ("L","R","T","B"))
    return (f'<a:tc><a:txBody><a:bodyPr wrap="square"/><a:lstStyle/><a:p><a:pPr algn="{algn}"/>{run}</a:p></a:txBody>'
            f'<a:tcPr marL="82296" marR="82296" marT="36576" marB="36576" anchor="ctr">{ln}<a:solidFill><a:srgbClr val="{fill}"/></a:solidFill></a:tcPr></a:tc>')

def _table(items, b):
    W=[560070,6050000,1750000,3010000]
    hdr=(f'<a:tr h="440000">{_tc("#",1100,True,"FFFFFF",b,"ctr")}{_tc("COMPROMISO (comité anterior)",1100,True,"FFFFFF",b,"l")}'
         f'{_tc("¿CUMPLIÓ?",1100,True,"FFFFFF",b,"ctr")}{_tc("COMENTARIOS",1100,True,"FFFFFF",b,"l")}</a:tr>')
    rows=hdr
    data=items if items else [""]*5     # vacío: 5 filas en blanco, mismo estilo
    N=len(data); szc=1050 if N<=5 else 950
    for i,it in enumerate(data):
        sh="F4F6F9" if i%2 else "FFFFFF"
        rows+=(f'<a:tr h="620000">{_tc(str(i+1),1200,True,"FFFFFF",b,"ctr")}{_tc(it,szc,False,"2B2B2B",sh,"l")}'
               f'{_tc("Sí  /  No",1000,False,"9AA0A6",sh,"ctr",True)}{_tc("",1000,False,"2B2B2B",sh,"l")}</a:tr>')
    grid="".join(f'<a:gridCol w="{w}"/>' for w in W)
    return (f'<p:graphicFrame><p:nvGraphicFramePr><p:cNvPr id="900" name="TablaCompromisos"/>'
            f'<p:cNvGraphicFramePr><a:graphicFrameLocks noGrp="1"/></p:cNvGraphicFramePr><p:nvPr/></p:nvGraphicFramePr>'
            f'<p:xfrm><a:off x="274638" y="1310000"/><a:ext cx="{sum(W)}" cy="500000"/></p:xfrm>'
            f'<a:graphic><a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/table">'
            f'<a:tbl><a:tblPr firstRow="1"/><a:tblGrid>{grid}</a:tblGrid>{rows}</a:tbl></a:graphicData></a:graphic></p:graphicFrame>')

def fill(pptx_bytes, xlsx_bytes):
    if not xlsx_bytes: return pptx_bytes
    comp=_read(xlsx_bytes)
    z=zipfile.ZipFile(io.BytesIO(pptx_bytes)); files={n:z.read(n) for n in z.namelist()}
    diag=[n for n in files if re.match(r'ppt/slides/slide\d+\.xml$',n) and "SEGUIMIENTO PLAN SEMANA" in files[n].decode("utf-8","ignore")]
    for path in diag:
        xml=files[path].decode()
        tt=[t for t in re.findall(r'<a:t>([^<]+)</a:t>',xml) if t.strip()]
        if not tt: continue
        s=_sid(tt[0]); items=comp.get(s) or (comp.get("EFECT") if s in ("DINERS","BANBIF") else None) or []
        d=minidom.parseString(xml); tree=d.getElementsByTagName("p:spTree")[0]
        for n in list(tree.childNodes):
            if n.nodeType!=1 or n.tagName not in ("p:sp","p:graphicFrame","p:pic"): continue
            off=n.getElementsByTagName("a:off")
            if not off or not off[0].getAttribute("x").strip(): continue
            y=int(off[0].getAttribute("y"))/E
            if 1.0<y<7.0: tree.removeChild(n)
        xml=d.toxml().replace("</p:spTree>", _table(items,_brand(s))+"</p:spTree>",1)
        files[path]=xml.encode()
    out=io.BytesIO()
    with zipfile.ZipFile(out,"w",zipfile.ZIP_DEFLATED) as zz:
        for n,dd in files.items(): zz.writestr(n,dd)
    return out.getvalue()
